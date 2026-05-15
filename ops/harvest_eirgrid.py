#!/usr/bin/env python3
"""
Best-effort EirGrid harvester for Ireland Energy Monitor.

Downloads EirGrid's quarter-hourly system data spreadsheet and attempts to extract:
- latest demand
- latest wind generation
- latest solar generation
- latest CO2 intensity
- net imports/interconnection, where detectable

This first harvester is deliberately conservative:
- it keeps the site working even if the spreadsheet layout changes
- it marks residual/gas as an estimate
- it writes source metadata and caveats
"""

from __future__ import annotations

import json
import math
import re
import sys
import urllib.request
from datetime import datetime, timezone, time, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
CACHE = ROOT / "ops" / "cache"
SOURCE = ROOT / "data" / "source"
ELECTRICITY_OUT = SOURCE / "electricity.json"
METADATA_OUT = SOURCE / "metadata.json"

EIRGRID_QTR_HOURLY_URL = (
    "https://cms.eirgrid.ie/sites/default/files/publications/"
    "System-Data-Qtr-Hourly-2026-V3.xlsx"
)


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(value):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_dt(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass

    return None


def download(url: str, dest: Path) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "IrelandEnergyTransitionMonitor/0.2 (+https://salmonofdoubt.github.io/demos/ietm/)"
        },
    )
    with urllib.request.urlopen(req, timeout=60) as response:
        dest.write_bytes(response.read())


def row_score(row: tuple[Any, ...]) -> int:
    joined = " ".join(norm(c) for c in row if c is not None)
    terms = [
        "date",
        "time",
        "demand",
        "wind",
        "solar",
        "co2",
        "carbon",
        "generation",
        "interconnector",
        "interconnection",
        "interchange",
    ]
    return sum(1 for term in terms if term in joined)


def find_header(sheet) -> tuple[int, list[str]] | None:
    best = None
    best_score = 0

    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
        score = row_score(row)
        if score > best_score:
            best_score = score
            best = (i, [str(c or "").strip() for c in row])

    if best and best_score >= 3:
        return best

    return None


def choose_col(headers: list[str], must: list[str], prefer: list[str] | None = None, avoid: list[str] | None = None) -> int | None:
    prefer = prefer or []
    avoid = avoid or []
    candidates = []

    for idx, header in enumerate(headers):
        h = norm(header)
        if not h:
            continue
        if not all(term in h for term in must):
            continue
        if any(term in h for term in avoid):
            continue

        score = 0
        for term in prefer:
            if term in h:
                score += 3
        if "roi" in h or "ireland" in h:
            score += 2
        if "northern" in h or h.startswith("ni"):
            score -= 4
        if "forecast" in h:
            score -= 2

        candidates.append((score, idx))

    if not candidates:
        return None

    candidates.sort(reverse=True)
    return candidates[0][1]


def extract_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)

    best_rows: list[dict[str, Any]] = []
    best_info: dict[str, Any] = {}

    for ws in wb.worksheets:
        header = find_header(ws)
        if not header:
            continue

        header_row, headers = header
        headers_norm = [norm(h) for h in headers]

        dt_col = (
            choose_col(headers, ["datetime"])
            or choose_col(headers, ["timestamp"])
            or choose_col(headers, ["date"])
        )
        time_col = choose_col(headers, ["time"])

        demand_col = choose_col(headers, ["demand"], prefer=["actual", "roi", "ireland"], avoid=["forecast"])
        wind_col = choose_col(headers, ["wind"], prefer=["actual", "generation", "roi", "ireland"], avoid=["forecast"])
        solar_col = choose_col(headers, ["solar"], prefer=["actual", "generation", "roi", "ireland"], avoid=["forecast"])
        co2_col = (
            choose_col(headers, ["co2"], prefer=["intensity", "gkwh", "gco2"])
            or choose_col(headers, ["carbon"], prefer=["intensity"])
        )
        import_col = (
            choose_col(headers, ["interconnection"], prefer=["net", "roi", "ireland"])
            or choose_col(headers, ["interconnector"], prefer=["net", "roi", "ireland"])
            or choose_col(headers, ["interchange"], prefer=["net", "roi", "ireland"])
            or choose_col(headers, ["import"], prefer=["net", "roi", "ireland"])
        )

        # Some EirGrid sheets expose interconnectors as separate EWIC/Moyle/Greenlink columns
        # rather than a single net-import column. Treat positive values as imports for now.
        interconnector_component_cols = []
        if import_col is None:
            for idx, header in enumerate(headers):
                h = norm(header)
                if any(term in h for term in ("ewic", "moyle", "greenlink")) and not any(
                    bad in h for bad in ("availability", "capacity", "forecast")
                ):
                    interconnector_component_cols.append(idx)

        if demand_col is None and wind_col is None and solar_col is None:
            continue

        rows: list[dict[str, Any]] = []

        for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            values = list(excel_row)

            def get(idx: int | None) -> Any:
                if idx is None or idx >= len(values):
                    return None
                return values[idx]

            dt = parse_dt(get(dt_col))
            if dt is None and dt_col is not None and time_col is not None:
                d = parse_dt(get(dt_col))
                t = get(time_col)
                if isinstance(t, time) and d:
                    dt = datetime.combine(d.date(), t)
                elif d and isinstance(t, datetime):
                    dt = datetime.combine(d.date(), t.time())

            demand = to_float(get(demand_col))
            wind = to_float(get(wind_col))
            solar = to_float(get(solar_col))
            co2 = to_float(get(co2_col))
            imports = to_float(get(import_col))

            if imports is None and interconnector_component_cols:
                component_values = [
                    to_float(get(idx))
                    for idx in interconnector_component_cols
                ]
                positive_imports = [
                    value for value in component_values
                    if value is not None and value > 0
                ]
                if positive_imports:
                    imports = sum(positive_imports)

            if demand is None and wind is None and solar is None:
                continue

            rows.append({
                "datetime": dt.isoformat() if dt else None,
                "demand_mw": demand,
                "wind_mw": wind,
                "solar_mw": solar,
                "co2_g_per_kwh": co2,
                "imports_mw": imports,
            })

        if len(rows) > len(best_rows):
            best_rows = rows
            best_info = {
                "sheet": ws.title,
                "header_row": header_row,
                "columns": {
                    "datetime": headers[dt_col] if dt_col is not None and dt_col < len(headers) else None,
                    "time": headers[time_col] if time_col is not None and time_col < len(headers) else None,
                    "demand": headers[demand_col] if demand_col is not None and demand_col < len(headers) else None,
                    "wind": headers[wind_col] if wind_col is not None and wind_col < len(headers) else None,
                    "solar": headers[solar_col] if solar_col is not None and solar_col < len(headers) else None,
                    "co2": headers[co2_col] if co2_col is not None and co2_col < len(headers) else None,
                    "imports": headers[import_col] if import_col is not None and import_col < len(headers) else None,
                    "interconnector_components": [
                        headers[idx] for idx in interconnector_component_cols
                        if idx < len(headers)
                    ],
                },
                "header_sample": headers[:20],
                "row_count": len(rows),
            }

    return best_rows, best_info


def average(values: list[float]) -> float | None:
    values = [v for v in values if v is not None and math.isfinite(v)]
    if not values:
        return None
    return sum(values) / len(values)


def pct(part: float | None, whole: float | None) -> float:
    if part is None or whole is None or whole <= 0:
        return 0.0
    return max(0.0, min(100.0, part / whole * 100.0))


def build_electricity(rows: list[dict[str, Any]], info: dict[str, Any]) -> dict[str, Any]:
    usable = [r for r in rows if r.get("demand_mw") or r.get("wind_mw") or r.get("solar_mw")]
    if not usable:
        raise RuntimeError("No usable rows found in EirGrid spreadsheet.")

    latest = usable[-1]
    window = usable[-96:] if len(usable) >= 96 else usable

    demand_now = latest.get("demand_mw") or average([r.get("demand_mw") for r in window]) or 0
    wind_now = latest.get("wind_mw") or 0
    solar_now = latest.get("solar_mw") or 0
    co2_values = [
        r.get("co2_g_per_kwh")
        for r in window
        if r.get("co2_g_per_kwh") is not None
    ]
    imports_values = [
        r.get("imports_mw")
        for r in window
        if r.get("imports_mw") is not None
    ]

    co2_available = bool(co2_values)
    imports_available = bool(imports_values)

    co2_now = latest.get("co2_g_per_kwh")
    if co2_now is None and co2_available:
        co2_now = average(co2_values)

    imports_now = latest.get("imports_mw")
    if imports_now is None:
        imports_now = 0

    wind_pct_now = pct(wind_now, demand_now)
    solar_pct_now = pct(solar_now, demand_now)
    imports_pct_now = pct(max(imports_now, 0), demand_now) if imports_available else 0
    renewables_pct_raw_now = max(0, wind_pct_now + solar_pct_now)
    renewables_pct_now = max(0, min(100, renewables_pct_raw_now))
    renewable_surplus_pct_now = max(0, renewables_pct_raw_now - 100)

    avg_demand = average([r.get("demand_mw") for r in window]) or demand_now
    avg_wind = average([r.get("wind_mw") for r in window]) or wind_now
    avg_solar = average([r.get("solar_mw") for r in window]) or solar_now
    avg_imports = average([max(v, 0) for v in imports_values]) if imports_available else 0

    wind_pct_24h = pct(avg_wind, avg_demand)
    solar_pct_24h = pct(avg_solar, avg_demand)
    imports_pct_24h = pct(avg_imports, avg_demand)

    # Thermal/other is not pure gas. It is all non-wind, non-solar, non-import supply.
    residual_pct_24h = max(0, min(100, 100 - wind_pct_24h - solar_pct_24h - imports_pct_24h))
    residual_pct_now = max(0, min(100, 100 - wind_pct_now - solar_pct_now - imports_pct_now))

    fuel_mix = [
        {"label": "Wind", "class": "wind", "percent": round(wind_pct_24h, 1), "available": True},
        {"label": "Solar", "class": "solar", "percent": round(solar_pct_24h, 1), "available": True},
        {"label": "Imports", "class": "imports", "percent": round(imports_pct_24h, 1), "available": imports_available},
        {"label": "Residual", "class": "other", "percent": round(residual_pct_24h, 1), "available": True},
    ]

    # Correct small rounding drift to keep validator happy.
    drift = round(100 - sum(item["percent"] for item in fuel_mix), 1)
    fuel_mix[-1]["percent"] = round(fuel_mix[-1]["percent"] + drift, 1)

    headline = (
        "Ireland is renewable-led in this quarter-hour signal."
        if renewables_pct_now >= residual_pct_now
        else "Ireland remains thermal/other-backed in this quarter-hour signal."
    )

    interpretation = (
        "This first live-linked module uses EirGrid quarter-hourly system data where column mapping is possible. "
        "Wind and solar are measured directly where available. Thermal/other is calculated as everything not "
        "identified here as wind, solar or net imports, so it should not yet be read as pure gas."
    )

    return {
        "electricity_now": {
            "demand_mw": round(demand_now),
            "renewables_percent": round(renewables_pct_now, 1),
            "renewables_output_percent": round(renewables_pct_raw_now, 1),
            "renewable_surplus_percent": round(renewable_surplus_pct_now, 1),
            "renewables_coverage_percent": round(renewables_pct_now, 1),
            "renewables_normalised": bool(renewable_surplus_pct_now > 0),
            "renewables_model": "eirgrid_workbook_wind_solar_cover_of_demand",
            "renewables_definition": "Wind plus solar as share of current demand from the EirGrid quarter-hourly workbook.",
            "wind_percent": round(wind_pct_now, 1),
            "solar_percent": round(solar_pct_now, 1),
            "gas_percent": round(residual_pct_now, 1),
            "residual_percent": round(residual_pct_now, 1),
            "imports_percent": round(imports_pct_now, 1),
            "imports_available": imports_available,
            "co2_g_per_kwh": round(co2_now, 1) if co2_available and co2_now is not None else None,
            "co2_available": co2_available,
            "gas_is_residual_proxy": True,
            "electricity_datetime": latest.get("datetime"),
            "source_label": "EirGrid quarter-hourly workbook",
            "source_url": EIRGRID_QTR_HOURLY_URL,
            "source_freshness": "latest mapped workbook interval",
            "data_age_hours": None,
        },
        "fuel_mix_24h": fuel_mix,
        "daily_story": {
            "headline": headline,
            "interpretation": interpretation,
        },
        "gas": {
            "share_percent": round(residual_pct_now, 1),
            "signal": "Thermal/other estimate",
            "narrative": (
                "Thermal/other is currently calculated as demand not covered by detected wind, solar and net imports. "
                "A later fuel-mix harvester should split this into gas, hydro, storage, coal/oil and other sources."
            ),
        },
        "source_status": {
            "source": "EirGrid System Data Qtr Hourly spreadsheet",
            "source_url": EIRGRID_QTR_HOURLY_URL,
            "harvested_at": now_iso(),
            "parser": info,
            "caveat": "Best-effort parser. Thermal/other is not pure gas."
        }
    }


def update_metadata(success: bool, message: str) -> None:
    existing = {}
    if METADATA_OUT.exists():
        try:
            existing = json.loads(METADATA_OUT.read_text())
        except json.JSONDecodeError:
            existing = {}

    existing.update({
        "project": "Ireland Energy Transition Monitor",
        "timezone": "Europe/Dublin",
        "mode": "Generated static dataset with EirGrid harvester",
        "confidence": "Medium" if success else "Low",
        "status": message,
        "last_eirgrid_harvest_at": now_iso(),
        "sources": [
            "EirGrid System Data Qtr Hourly spreadsheet",
            "SEAI energy statistics and prices, planned",
            "SEAI renewable electricity county dashboard, planned",
            "CSO transport indicators, planned",
            "Gas Networks Ireland gas demand reporting, planned"
        ]
    })

    METADATA_OUT.write_text(json.dumps(existing, indent=2) + "\n")


def main() -> int:
    CACHE.mkdir(parents=True, exist_ok=True)
    SOURCE.mkdir(parents=True, exist_ok=True)

    xlsx_path = CACHE / "eirgrid_system_data_qtr_hourly.xlsx"

    try:
        print(f"Downloading {EIRGRID_QTR_HOURLY_URL}")
        download(EIRGRID_QTR_HOURLY_URL, xlsx_path)

        rows, info = extract_rows(xlsx_path)
        electricity = build_electricity(rows, info)

        ELECTRICITY_OUT.write_text(json.dumps(electricity, indent=2) + "\n")
        update_metadata(True, "EirGrid harvester succeeded; live-linked electricity source active.")

        print(f"Wrote {ELECTRICITY_OUT.relative_to(ROOT)}")
        print(f"Rows parsed: {len(rows)}")
        print(f"Parser info: {json.dumps(info, indent=2)[:1000]}")
        return 0

    except Exception as exc:
        message = f"EirGrid harvester failed; keeping previous electricity source. Reason: {exc}"
        print(f"WARNING: {message}", file=sys.stderr)
        update_metadata(False, message)
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
